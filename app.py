import streamlit as st
import pandas as pd
import time
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from rapidfuzz import fuzz
import os
from openpyxl import load_workbook
from langchain.prompts import PromptTemplate
from langchain_core.runnables import RunnablePassthrough
from io import StringIO, BytesIO
import sys
import contextlib
from langchain_openai import ChatOpenAI  # Updated import
import pdfkit
from jinja2 import Template
import time
from tenacity import retry, stop_after_attempt, wait_exponential
from typing import Optional
import torch
from transformers import (
    pipeline,
    AutoModelForSeq2SeqLM,
    AutoTokenizer,
    AutoModelForCausalLM  # 4 Qwen
)

from threading import Event
import threading
from queue import Queue

from deep_translator import GoogleTranslator
from googletrans import Translator as LegacyTranslator
import plotly.graph_objects as go
from datetime import datetime
import plotly.express as px


class ProcessControl:
    def __init__(self):
        self.pause_event = Event()
        self.stop_event = Event()
        self.pause_event.set()  # Start in non-paused state
        
    def pause(self):
        self.pause_event.clear()
        
    def resume(self):
        self.pause_event.set()
        
    def stop(self):
        self.stop_event.set()
        self.pause_event.set()  # Ensure not stuck in pause
        
    def reset(self):
        self.stop_event.clear()
        self.pause_event.set()
        
    def is_paused(self):
        return not self.pause_event.is_set()
        
    def is_stopped(self):
        return self.stop_event.is_set()
        
    def wait_if_paused(self):
        self.pause_event.wait()


class FallbackLLMSystem:
    def __init__(self):
        """Initialize fallback models for event detection and reasoning"""
        try:
            # Initialize MT5 model (multilingual T5)
            self.model_name = "google/mt5-small"
            self.tokenizer = AutoTokenizer.from_pretrained(self.model_name)
            self.model = AutoModelForSeq2SeqLM.from_pretrained(self.model_name)
            
            # Set device
            self.device = "cuda" if torch.cuda.is_available() else "cpu"
            self.model = self.model.to(self.device)
            
            st.success(f"пока все в порядке: запущена MT5 model на = {self.device} =")
            
        except Exception as e:
            st.error(f"Ошибка запуска модели MT5: {str(e)}")
            raise

    def invoke(self, prompt_args):
        """Make the class compatible with LangChain by implementing invoke"""
        try:
            if isinstance(prompt_args, dict):
                # Extract the prompt template result
                template_result = prompt_args.get('template_result', '')
                if not template_result:
                    # Try to construct from entity and news if available
                    entity = prompt_args.get('entity', '')
                    news = prompt_args.get('news', '')
                    template_result = f"Analyze news about {entity}: {news}"
            else:
                template_result = str(prompt_args)

            # Process with MT5
            inputs = self.tokenizer(
                template_result,
                return_tensors="pt",
                padding=True,
                truncation=True,
                max_length=512
            ).to(self.device)
            
            outputs = self.model.generate(
                **inputs,
                max_length=200,
                num_return_sequences=1,
                do_sample=False,
                pad_token_id=self.tokenizer.pad_token_id
            )
            
            response = self.tokenizer.decode(outputs[0], skip_special_tokens=True)
            
            # Return in a format compatible with LangChain
            return type('Response', (), {'content': response})()
            
        except Exception as e:
            st.warning(f"MT5 generation error: {str(e)}")
            # Return a default response on error
            return type('Response', (), {
                'content': 'Impact: Неопределенный эффект\nReasoning: Ошибка анализа'
            })()

    def __or__(self, other):
        """Implement the | operator for chain compatibility"""
        if callable(other):
            return lambda x: other(self(x))
        return NotImplemented

    def __rrshift__(self, other):
        """Implement the >> operator for chain compatibility"""
        return self.__or__(other)

    def __call__(self, prompt_args):
        """Make the class callable for chain compatibility"""
        return self.invoke(prompt_args)

    def detect_events(self, text: str, entity: str) -> tuple[str, str]:
        """
        Detect events using MT5 with improved error handling and response parsing
        
        Args:
            text (str): The news text to analyze
            entity (str): The company/entity name
            
        Returns:
            tuple[str, str]: (event_type, summary)
        """
        # Initialize default return values
        event_type = "Нет"
        summary = ""
        
        # Input validation
        if not text or not entity or not isinstance(text, str) or not isinstance(entity, str):
            return event_type, "Invalid input"
            
        try:
            # Clean and prepare input text
            text = text.strip()
            entity = entity.strip()
            
            # Construct prompt with better formatting
            prompt = f"""<s>Analyze the following news about {entity}:

    Text: {text}

    Task: Identify the main event type and provide a brief summary.

    Event types:
    1. Отчетность - Events related to financial reports, earnings, revenue, EBITDA
    2. РЦБ - Events related to securities, bonds, stock market, defaults, restructuring
    3. Суд - Events related to legal proceedings, lawsuits, arbitration
    4. Нет - No significant events detected

    Required output format:
    Тип: [event type]
    Краткое описание: [1-2 sentence summary]</s>"""

            # Process with MT5
            try:
                inputs = self.tokenizer(
                    prompt,
                    return_tensors="pt",
                    padding=True,
                    truncation=True,
                    max_length=512
                ).to(self.device)
                
                outputs = self.model.generate(
                    **inputs,
                    max_length=300,  # Increased for better summaries
                    num_return_sequences=1,
                    do_sample=False,
                    pad_token_id=self.tokenizer.pad_token_id,
                    eos_token_id=self.tokenizer.eos_token_id,
                    no_repeat_ngram_size=3  # Prevent repetition
                )
                
                response = self.tokenizer.decode(outputs[0], skip_special_tokens=True)
                
            except torch.cuda.OutOfMemoryError:
                st.warning("GPU memory exceeded, falling back to CPU")
                self.model = self.model.to('cpu')
                inputs = inputs.to('cpu')
                outputs = self.model.generate(
                    **inputs,
                    max_length=300,
                    num_return_sequences=1,
                    do_sample=False,
                    pad_token_id=self.tokenizer.pad_token_id
                )
                response = self.tokenizer.decode(outputs[0], skip_special_tokens=True)
                self.model = self.model.to(self.device)  # Move back to GPU
                
            # Enhanced response parsing
            if "Тип:" in response and "Краткое описание:" in response:
                try:
                    # Split and clean parts
                    parts = response.split("Краткое описание:")
                    type_part = parts[0].split("Тип:")[1].strip()
                    
                    # Validate event type with fuzzy matching
                    valid_types = ["Отчетность", "РЦБ", "Суд", "Нет"]
                    
                    # Check for exact matches first
                    if type_part in valid_types:
                        event_type = type_part
                    else:
                        # Check keywords for each type
                        keywords = {
                            "Отчетность": ["отчет", "выручка", "прибыль", "ebitda", "финанс"],
                            "РЦБ": ["облигаци", "купон", "дефолт", "реструктуризац", "ценные бумаги"],
                            "Суд": ["суд", "иск", "арбитраж", "разбирательств"]
                        }
                        
                        # Look for keywords in both type and summary
                        full_text = response.lower()
                        for event_category, category_keywords in keywords.items():
                            if any(keyword in full_text for keyword in category_keywords):
                                event_type = event_category
                                break
                    
                    # Extract and clean summary
                    if len(parts) > 1:
                        summary = parts[1].strip()
                        # Ensure summary isn't too long
                        if len(summary) > 200:
                            summary = summary[:197] + "..."
                        
                        # Add entity reference if missing
                        if entity.lower() not in summary.lower():
                            summary = f"Компания {entity}: {summary}"
                    
                except IndexError:
                    st.warning("Error parsing model response format")
                    return "Нет", "Error parsing response"
                    
            # Additional validation
            if not summary or len(summary) < 5:
                keywords = {
                    "Отчетность": "Обнаружена информация о финансовой отчетности",
                    "РЦБ": "Обнаружена информация о ценных бумагах",
                    "Суд": "Обнаружена информация о судебном разбирательстве",
                    "Нет": "Значимых событий не обнаружено"
                }
                summary = f"{keywords.get(event_type, 'Требуется дополнительный анализ')} ({entity})"
                
            return event_type, summary
                
        except Exception as e:
            st.warning(f"Event detection error: {str(e)}")
            # Try to provide more specific error information
            if "CUDA" in str(e):
                return "Нет", "GPU error - falling back to CPU needed"
            elif "tokenizer" in str(e):
                return "Нет", "Text processing error"
            elif "model" in str(e):
                return "Нет", "Model inference error"
            else:
                return "Нет", "Ошибка анализа"
        

def ensure_groq_llm():
    """Initialize Groq LLM for impact estimation"""
    try:
        if 'groq_key' not in st.secrets:
            st.error("Groq API key not found in secrets. Please add it with the key 'groq_key'.")
            return None
            
        return ChatOpenAI(
            base_url="https://api.groq.com/openai/v1",
            model="llama-3.1-70b-versatile",
            openai_api_key=st.secrets['groq_key'],
            temperature=0.0
        )
    except Exception as e:
        st.error(f"Error initializing Groq LLM: {str(e)}")
        return None

def estimate_impact(llm, news_text, entity):
    """
    Estimate impact using Groq LLM regardless of the main model choice.
    Falls back to the provided LLM if Groq initialization fails.
    """
    # Initialize default return values
    impact = "Неопределенный эффект"
    reasoning = "Не удалось получить обоснование"
    
    try:
        # Always try to use Groq first
        groq_llm = ensure_groq_llm()
        working_llm = groq_llm if groq_llm is not None else llm
        
        template = """
        You are a financial analyst. Analyze this news piece about {entity} and assess its potential impact.
        
        News: {news}
        
        Classify the impact into one of these categories:
        1. "Значительный риск убытков" (Significant loss risk)
        2. "Умеренный риск убытков" (Moderate loss risk)
        3. "Незначительный риск убытков" (Minor loss risk)
        4. "Вероятность прибыли" (Potential profit)
        5. "Неопределенный эффект" (Uncertain effect)
        
        Provide a brief, fact-based reasoning for your assessment.
        
        Format your response exactly as:
        Impact: [category]
        Reasoning: [explanation in 2-3 sentences]
        """
        
        prompt = PromptTemplate(template=template, input_variables=["entity", "news"])
        chain = prompt | working_llm
        response = chain.invoke({"entity": entity, "news": news_text})
        
        # Extract content from response
        response_text = response.content if hasattr(response, 'content') else str(response)
        
        if "Impact:" in response_text and "Reasoning:" in response_text:
            impact_part, reasoning_part = response_text.split("Reasoning:")
            impact_temp = impact_part.split("Impact:")[1].strip()
            
            # Validate impact category
            valid_impacts = [
                "Значительный риск убытков",
                "Умеренный риск убытков",
                "Незначительный риск убытков",
                "Вероятность прибыли",
                "Неопределенный эффект"
            ]
            if impact_temp in valid_impacts:
                impact = impact_temp
            reasoning = reasoning_part.strip()
            
    except Exception as e:
        st.warning(f"Error in impact estimation: {str(e)}")
    
    return impact, reasoning
   
class QwenSystem:
    def __init__(self):
        """Initialize Qwen 2.5 Coder model"""
        try:
            self.model_name = "Qwen/Qwen2.5-Coder-32B-Instruct"
            
            # Initialize model with auto settings
            self.model = AutoModelForCausalLM.from_pretrained(
                self.model_name,
                torch_dtype="auto",
                device_map="auto"
            )
            self.tokenizer = AutoTokenizer.from_pretrained(self.model_name)
            
            st.success(f"запустил Qwen2.5 model")
            
        except Exception as e:
            st.error(f"ошибка запуска Qwen2.5: {str(e)}")
            raise

    def invoke(self, messages):
        """Process messages using Qwen's chat template"""
        try:
            # Prepare messages with system prompt
            chat_messages = [
                {"role": "system", "content": "You are wise financial analyst. You are a helpful assistant."}
            ]
            chat_messages.extend(messages)
            
            # Apply chat template
            text = self.tokenizer.apply_chat_template(
                chat_messages,
                tokenize=False,
                add_generation_prompt=True
            )
            
            # Prepare model inputs
            model_inputs = self.tokenizer([text], return_tensors="pt").to(self.model.device)
            
            # Generate response
            generated_ids = self.model.generate(
                **model_inputs,
                max_new_tokens=512,
                pad_token_id=self.tokenizer.pad_token_id,
                eos_token_id=self.tokenizer.eos_token_id
            )
            
            # Extract new tokens
            generated_ids = [
                output_ids[len(input_ids):] 
                for input_ids, output_ids in zip(model_inputs.input_ids, generated_ids)
            ]
            
            # Decode response
            response = self.tokenizer.batch_decode(generated_ids, skip_special_tokens=True)[0]
            
            # Return in ChatOpenAI-compatible format
            return type('Response', (), {'content': response})()
            
        except Exception as e:
            st.warning(f"Qwen generation error: {str(e)}")
            raise


class ProcessingUI:
    def __init__(self):
        if 'control' not in st.session_state:
            st.session_state.control = ProcessControl()
            
        # Initialize processing stats in session state if not exists
        if 'processing_stats' not in st.session_state:
            st.session_state.processing_stats = {
                'start_time': time.time(),
                'entities': {},
                'events_timeline': [],
                'negative_alerts': [],
                'processing_speed': []
            }
        
        # Create main layout
        self.setup_layout()
        
    def setup_layout(self):
        """Setup the main UI layout with tabs and sections"""
        # Control Panel
        with st.container():
            col1, col2, col3 = st.columns([2,2,1])
            with col1:
                if st.button(
                    "⏸️ Пауза" if not st.session_state.control.is_paused() else "▶️ Продолжить",
                    use_container_width=True
                ):
                    if st.session_state.control.is_paused():
                        st.session_state.control.resume()
                    else:
                        st.session_state.control.pause()
            with col2:
                if st.button("⏹️ Остановить", use_container_width=True):
                    st.session_state.control.stop()
            with col3:
                self.timer_display = st.empty()
        
        # Progress Bar with custom styling
        st.markdown("""
            <style>
            .stProgress > div > div > div > div {
                background-image: linear-gradient(to right, #FF6B6B, #4ECDC4);
            }
            </style>""", 
            unsafe_allow_html=True
        )
        self.progress_bar = st.progress(0)
        self.status = st.empty()

        # Create tabs for different views
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Основные метрики", 
            "🏢 По организациям", 
            "⚠️ Важные события", 
            "📈 Аналитика"
        ])
        
        with tab1:
            self.setup_main_metrics_tab()
            
        with tab2:
            self.setup_entity_tab()
            
        with tab3:
            self.setup_events_tab()
            
        with tab4:
            self.setup_analytics_tab()
                  
    def setup_entity_tab(self):
        """Setup the entity-wise analysis display"""
        # Entity filter
        self.entity_filter = st.multiselect(
            "Фильтр по организациям:",
            options=[],  # Will be populated as entities are processed
            default=None
        )
        
        # Entity metrics
        self.entity_cols = st.columns([2,1,1,1])
        self.entity_chart = st.empty()
        self.entity_table = st.empty()
        
    def setup_events_tab(self):
        """Setup the events timeline display"""
        # Event type filter - store in session state
        if 'event_filter' not in st.session_state:
            st.session_state.event_filter = []
            
        st.session_state.event_filter = st.multiselect(
            "Тип события:",
            options=["Отчетность", "РЦБ", "Суд"],
            default=None,
            key="event_filter_key"
        )
        
        self.timeline_container = st.container()
    
    def _update_events_view(self, row, event_type):
        """Update events timeline"""
        if event_type != 'Нет':
            event_html = f"""
                <div class='timeline-item' style='
                    border-left: 4px solid #2196F3;
                    margin: 10px 0;
                    padding: 10px;
                    background: #f5f5f5;
                    border-radius: 4px;
                '>
                    <h4 style='color: #2196F3; margin: 0;'>{event_type}</h4>
                    <p><strong>{row['Объект']}</strong></p>
                    <p>{row['Заголовок']}</p>
                    <p style='font-size: 0.9em;'>{row['Выдержки из текста']}</p>
                    <small style='color: #666;'>{datetime.now().strftime('%H:%M:%S')}</small>
                </div>
            """
            with self.timeline_container:
                st.markdown(event_html, unsafe_allow_html=True)

    def setup_analytics_tab(self):
        """Setup the analytics display"""
        # Create containers for analytics
        self.speed_container = st.container()
        with self.speed_container:
            st.subheader("Скорость обработки")
            self.speed_chart = st.empty()
            
        self.sentiment_container = st.container()
        with self.sentiment_container:
            st.subheader("Распределение тональности")
            self.sentiment_chart = st.empty()
            
        self.correlation_container = st.container()
        with self.correlation_container:
            st.subheader("Корреляция между метриками")
            self.correlation_chart = st.empty()
        
    def update_stats(self, row, sentiment, event_type, processing_speed):
        """Update all statistics and displays"""
        # Update session state stats
        stats = st.session_state.processing_stats
        entity = row['Объект']
        
        # Update entity stats
        if entity not in stats['entities']:
            stats['entities'][entity] = {
                'total': 0,
                'negative': 0,
                'events': 0,
                'timeline': []
            }
        
        stats['entities'][entity]['total'] += 1
        if sentiment == 'Negative':
            stats['entities'][entity]['negative'] += 1
        if event_type != 'Нет':
            stats['entities'][entity]['events'] += 1
            
        # Update processing speed
        stats['processing_speed'].append(processing_speed)
        
        # Update UI components
        self._update_main_metrics(row, sentiment, event_type, processing_speed)
        self._update_entity_view()
        self._update_events_view(row, event_type)
        self._update_analytics()
        
    def _update_main_metrics(self, row, sentiment, event_type, speed):
        """Update main metrics tab"""
        total = sum(e['total'] for e in st.session_state.processing_stats['entities'].values())
        total_negative = sum(e['negative'] for e in st.session_state.processing_stats['entities'].values())
        total_events = sum(e['events'] for e in st.session_state.processing_stats['entities'].values())
        
        # Update metrics
        self.total_processed.metric("Обработано", total)
        self.negative_count.metric("Негативных", total_negative)
        self.events_count.metric("Событий", total_events)
        self.speed_metric.metric("Скорость", f"{speed:.1f} сообщ/сек")
        
        # Update recent items
        self._update_recent_items(row, sentiment, event_type)
        
    def _update_recent_items(self, row, sentiment, event_type):
        """Update recent items display using Streamlit native components"""
        if 'recent_items' not in st.session_state:
            st.session_state.recent_items = []
            
        # Add new item to the list
        new_item = {
            'entity': row['Объект'],
            'headline': row['Заголовок'],
            'sentiment': sentiment,
            'event_type': event_type,
            'time': datetime.now().strftime('%H:%M:%S')
        }
        
        # Update the list in session state
        if not any(
            item['entity'] == new_item['entity'] and 
            item['headline'] == new_item['headline'] 
            for item in st.session_state.recent_items
        ):
            st.session_state.recent_items.insert(0, new_item)
            st.session_state.recent_items = st.session_state.recent_items[:10]  # Keep last 10 items

        # Prepare markdown for all items
        all_items_markdown = ""
        
        for item in st.session_state.recent_items:
            if item['sentiment'] in ['Positive', 'Negative']:
                sentiment_color = "🔴" if item['sentiment'] == 'Negative' else "🟢"
                event_icon = "📅" if item['event_type'] != 'Нет' else ""
                
                event_text = f" | Событие: {item['event_type']}" if item['event_type'] != 'Нет' else ""
                
                all_items_markdown += f"""
                {sentiment_color} **{item['entity']}**  {event_icon}

                {item['headline']}

                *{item['sentiment']}*{event_text} | {item['time']}

                ---
                """
        
        # Update container with all items at once
        if all_items_markdown:
            self.recent_items_container.markdown(all_items_markdown)

    def setup_main_metrics_tab(self):
        """Setup the main metrics display with updated styling"""
        # Create metrics containers
        metrics_cols = st.columns(4)
        self.total_processed = metrics_cols[0].empty()
        self.negative_count = metrics_cols[1].empty()
        self.events_count = metrics_cols[2].empty()
        self.speed_metric = metrics_cols[3].empty()
        
        # Create container for recent items
        st.markdown("### негативные/позитивные")
        self.recent_items_container = st.empty()


    def _update_entity_view(self):
        """Update entity tab visualizations"""
        stats = st.session_state.processing_stats['entities']
        if not stats:
            return
            
        # Get filtered entities
        filtered_entities = self.entity_filter or stats.keys()
        
        # Create entity comparison chart using Plotly
        df_entities = pd.DataFrame.from_dict(stats, orient='index')
        df_entities = df_entities.loc[filtered_entities]  # Apply filter
        
        fig = go.Figure(data=[
            go.Bar(
                name='Всего',
                x=df_entities.index,
                y=df_entities['total'],
                marker_color='#E0E0E0'  # Light gray
            ),
            go.Bar(
                name='Негативные',
                x=df_entities.index,
                y=df_entities['negative'],
                marker_color='#FF6B6B'  # Red
            ),
            go.Bar(
                name='События',
                x=df_entities.index,
                y=df_entities['events'],
                marker_color='#2196F3'  # Blue
            )
        ])
        
        fig.update_layout(
            barmode='group',
            title='Статистика по организациям',
            xaxis_title='Организация',
            yaxis_title='Количество',
            showlegend=True
        )
        
        self.entity_chart.plotly_chart(fig, use_container_width=True)
                
    def _update_analytics(self):
        """Update analytics tab visualizations"""
        stats = st.session_state.processing_stats
        
        # Processing speed chart - showing last 20 measurements
        speeds = stats['processing_speed'][-20:]
        if speeds:
            fig_speed = go.Figure(data=go.Scatter(
                y=speeds,
                mode='lines+markers',
                name='Скорость',
                line=dict(color='#4CAF50')
            ))
            fig_speed.update_layout(
                title='Скорость обработки',
                yaxis_title='Сообщений в секунду',
                showlegend=True
            )
            self.speed_chart.plotly_chart(fig_speed, use_container_width=True)
            
        # Sentiment distribution pie chart
        if stats['entities']:
            total_negative = sum(e['negative'] for e in stats['entities'].values())
            total_positive = sum(e['events'] for e in stats['entities'].values())
            total_neutral = sum(e['total'] for e in stats['entities'].values()) - total_negative - total_positive
            
            fig_sentiment = go.Figure(data=[go.Pie(
                labels=['Негативные', 'Позитивные', 'Нейтральные'],
                values=[total_negative, total_positive, total_neutral],
                marker_colors=['#FF6B6B', '#4ECDC4', '#95A5A6']
            )])
            self.sentiment_chart.plotly_chart(fig_sentiment, use_container_width=True)
        
    def update_progress(self, current, total):
        """Update progress bar, elapsed time and estimated time remaining"""
        progress = current / total
        self.progress_bar.progress(progress)
        self.status.text(f"Обрабатываем {current} из {total} сообщений...")
        
        # Calculate times
        current_time = time.time()
        elapsed = current_time - st.session_state.processing_stats['start_time']
        
        # Calculate processing speed and estimated time remaining
        if current > 0:
            speed = current / elapsed  # items per second
            remaining_items = total - current
            estimated_remaining = remaining_items / speed if speed > 0 else 0
            
            time_display = (
                f"⏱️ Прошло: {format_elapsed_time(elapsed)} | "
                f"Осталось: {format_elapsed_time(estimated_remaining)}"
            )
        else:
            time_display = f"⏱️ Прошло: {format_elapsed_time(elapsed)}"
            
        self.timer_display.markdown(time_display)


class EventDetectionSystem:
    def __init__(self):
        try:
            # Initialize models with specific labels
            self.finbert = pipeline(
                "text-classification", 
                model="ProsusAI/finbert",
                return_all_scores=True
            )
            self.business_classifier = pipeline(
                "text-classification", 
                model="yiyanghkust/finbert-tone",
                return_all_scores=True
            )
            st.success("продолжается пока хорошо: BERT-модели запущены для детекции новостей")
        except Exception as e:
            st.error(f"Ошибка запуска BERT: {str(e)}")
            raise

    def detect_event_type(self, text, entity):
        event_type = "Нет"
        summary = ""
        
        try:
            # Ensure text is properly formatted
            text = str(text).strip()
            if not text:
                return "Нет", "Empty text"

            # Get predictions
            finbert_scores = self.finbert(
                text,
                truncation=True,
                max_length=512
            )
            business_scores = self.business_classifier(
                text,
                truncation=True,
                max_length=512
            )
            
            # Get highest scoring predictions
            finbert_pred = max(finbert_scores[0], key=lambda x: x['score'])
            business_pred = max(business_scores[0], key=lambda x: x['score'])
            
            # Map to event types with confidence threshold
            confidence_threshold = 0.6
            max_confidence = max(finbert_pred['score'], business_pred['score'])
            
            if max_confidence >= confidence_threshold:
                if any(term in text.lower() for term in ['отчет', 'выручка', 'прибыль', 'ebitda']):
                    event_type = "Отчетность"
                    summary = f"Финансовая отчетность (confidence: {max_confidence:.2f})"
                elif any(term in text.lower() for term in ['облигаци', 'купон', 'дефолт', 'реструктуризац']):
                    event_type = "РЦБ"
                    summary = f"Событие РЦБ (confidence: {max_confidence:.2f})"
                elif any(term in text.lower() for term in ['суд', 'иск', 'арбитраж']):
                    event_type = "Суд"
                    summary = f"Судебное разбирательство (confidence: {max_confidence:.2f})"
            
            if event_type != "Нет":
                summary += f"\nКомпания: {entity}"
            
            return event_type, summary
            
        except Exception as e:
            st.warning(f"Event detection error: {str(e)}")
            return "Нет", "Error in event detection"

class TranslationSystem:
    def __init__(self):
        """Initialize translation system using Helsinki NLP model with fallback options"""
        try:
            self.translator = pipeline("translation", model="Helsinki-NLP/opus-mt-ru-en")
            # Initialize fallback translator
            self.fallback_translator = GoogleTranslator(source='ru', target='en')
            self.legacy_translator = LegacyTranslator()
            st.success("начинается все хорошо: запустил систему перевода")
        except Exception as e:
            st.error(f"Ошибка запуска перевода: {str(e)}")
            raise

    def _split_into_chunks(self, text: str, max_length: int = 450) -> list:
        """Split text into chunks while preserving word boundaries"""
        words = text.split()
        chunks = []
        current_chunk = []
        current_length = 0

        for word in words:
            word_length = len(word)
            if current_length + word_length + 1 <= max_length:
                current_chunk.append(word)
                current_length += word_length + 1
            else:
                if current_chunk:
                    chunks.append(' '.join(current_chunk))
                current_chunk = [word]
                current_length = word_length

        if current_chunk:
            chunks.append(' '.join(current_chunk))

        return chunks

    def _translate_chunk_with_retries(self, chunk: str, max_retries: int = 3) -> str:
        """Attempt translation with multiple fallback options"""
        if not chunk or not chunk.strip():
            return ""

        for attempt in range(max_retries):
            try:
                # First try Helsinki NLP
                result = self.translator(chunk, max_length=512)
                if result and isinstance(result, list) and len(result) > 0:
                    translated = result[0].get('translation_text')
                    if translated and isinstance(translated, str):
                        return translated

                # First fallback: Google Translator
                translated = self.fallback_translator.translate(chunk)
                if translated and isinstance(translated, str):
                    return translated

                # Second fallback: Legacy Google Translator
                translated = self.legacy_translator.translate(chunk, src='ru', dest='en').text
                if translated and isinstance(translated, str):
                    return translated

            except Exception as e:
                if attempt == max_retries - 1:
                    st.warning(f"Попробовал перевести {max_retries} раз, не преуспел: {str(e)}")
                time.sleep(1 * (attempt + 1))  # Exponential backoff

        return chunk  # Return original text if all translation attempts fail

    def translate_text(self, text: str) -> str:
        """Translate text with robust error handling and validation"""
        # Input validation
        if pd.isna(text) or not isinstance(text, str):
            return str(text) if pd.notna(text) else ""

        text = str(text).strip()
        if not text:
            return ""

        try:
            # Split into manageable chunks
            chunks = self._split_into_chunks(text)
            translated_chunks = []

            # Process each chunk with validation
            for chunk in chunks:
                if not chunk.strip():
                    continue

                translated_chunk = self._translate_chunk_with_retries(chunk)
                if translated_chunk:  # Only add non-empty translations
                    translated_chunks.append(translated_chunk)
                time.sleep(0.1)  # Rate limiting

            # Final validation of results
            if not translated_chunks:
                return text  # Return original if no translations succeeded

            result = ' '.join(translated_chunks)
            return result if result.strip() else text

        except Exception as e:
            st.warning(f"Translation error: {str(e)}")
            return text  # Return original text on error



def process_file(uploaded_file, model_choice, translation_method=None):
    df = None
    processed_rows_df = pd.DataFrame()
    last_update_time = time.time()

    try:
        # Initialize UI and control systems
        ui = ProcessingUI()
        translator = TranslationSystem()
        event_detector = EventDetectionSystem()
        
        # Load and prepare data
        df = pd.read_excel(uploaded_file, sheet_name='Публикации')
        llm = init_langchain_llm(model_choice)
        
        # Initialize Groq for impact estimation
        groq_llm = ensure_groq_llm()
        if groq_llm is None:
            st.warning("Failed to initialize Groq LLM for impact estimation. Using fallback model.")
        
        # Initialize all required columns at the start
        required_columns = {
            'Объект': '',
            'Заголовок': '',
            'Выдержки из текста': '',
            'Translated': '',
            'Sentiment': 'Neutral',
            'Impact': 'Неопределенный эффект',
            'Reasoning': 'Не проанализировано',
            'Event_Type': 'Нет',
            'Event_Summary': ''
        }
        
        # Ensure all required columns exist in DataFrame
        for col, default_value in required_columns.items():
            if col not in df.columns:
                df[col] = default_value
        
        # Create processed_rows_df with all columns from original df and required columns
        all_columns = list(set(list(df.columns) + list(required_columns.keys())))
        processed_rows_df = pd.DataFrame(columns=all_columns)
        
        # Process rows
        total_rows = len(df)
        processed_rows = 0
        
        for idx, row in df.iterrows():
            if st.session_state.control.is_stopped():
                st.warning("Обработку остановили")
                if not processed_rows_df.empty:
                    try:
                        # Create the output files for each sheet
                        monitoring_df = processed_rows_df[processed_rows_df['Event_Type'] != 'Нет'].copy()
                        svodka_df = processed_rows_df.groupby('Объект').agg({
                            'Объект': 'first',
                            'Sentiment': lambda x: sum(x == 'Negative'),
                            'Event_Type': lambda x: sum(x != 'Нет')
                        }).reset_index()
                        
                        # Prepare final DataFrame for file creation
                        result_df = pd.DataFrame()
                        result_df['Мониторинг'] = monitoring_df.to_dict('records')
                        result_df['Сводка'] = svodka_df.to_dict('records')
                        result_df['Публикации'] = processed_rows_df.to_dict('records')
                        
                        output = create_output_file(result_df, uploaded_file, llm)
                        if output is not None:
                            st.download_button(
                                label=f"📊 Скачать результат ({processed_rows} из {total_rows} строк)",
                                data=output,
                                file_name="partial_analysis.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="partial_download"
                            )
                    except Exception as e:
                        st.error(f"Ошибка при создании файла: {str(e)}")
                        
                return processed_rows_df
                
            st.session_state.control.wait_if_paused()
            if st.session_state.control.is_paused():
                continue
                
            try:
                # Copy original row data
                new_row = row.copy()
                
                # Translation
                translated_text = translator.translate_text(row['Выдержки из текста'])
                new_row['Translated'] = translated_text
                
                # Sentiment analysis
                sentiment = analyze_sentiment(translated_text)
                new_row['Sentiment'] = sentiment
                
                # Event detection
                event_type, event_summary = event_detector.detect_event_type(
                    row['Выдержки из текста'],
                    row['Объект']
                )
                new_row['Event_Type'] = event_type
                new_row['Event_Summary'] = event_summary
                
                # Handle negative sentiment
                if sentiment == "Negative":
                    try:
                        if translated_text and len(translated_text.strip()) > 0:
                            impact, reasoning = estimate_impact(
                                groq_llm if groq_llm is not None else llm,
                                translated_text,
                                row['Объект']
                            )
                            new_row['Impact'] = impact
                            new_row['Reasoning'] = reasoning
                    except Exception as e:
                        new_row['Impact'] = "Неопределенный эффект"
                        new_row['Reasoning'] = "Ошибка анализа"
                
                # Add processed row to DataFrame
                processed_rows_df = pd.concat([processed_rows_df, pd.DataFrame([new_row])], ignore_index=True)
                
                # Update progress
                processed_rows += 1
                ui.update_progress(processed_rows, total_rows)
                
            except Exception as e:
                st.warning(f"Ошибка в обработке ряда {idx + 1}: {str(e)}")
                continue
                
        return processed_rows_df
        
    except Exception as e:
        st.error(f"Ошибка в обработке файла: {str(e)}")
        return None


    

def create_download_section(excel_data, pdf_data):
    st.markdown("""
        <div class="download-container">
            <div class="download-header">📥 Результаты анализа доступны для скачивания:</div>
        </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    
    with col1:
        if excel_data is not None:
            st.download_button(
                label="📊 Скачать Excel отчет",
                data=excel_data,
                file_name="результат_анализа.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="excel_download"
            )
        else:
            st.error("Ошибка при создании Excel файла")
    



def display_sentiment_results(row, sentiment, impact=None, reasoning=None):
    if sentiment == "Negative":
        st.markdown(f"""
            <div style='color: red; font-weight: bold;'>
            Объект: {row['Объект']}<br>
            Новость: {row['Заголовок']}<br>
            Тональность: {sentiment}<br>
            {"Эффект: " + impact + "<br>" if impact else ""}
            {"Обоснование: " + reasoning + "<br>" if reasoning else ""}
            </div>
            """, unsafe_allow_html=True)
    elif sentiment == "Positive":
        st.markdown(f"""
            <div style='color: green; font-weight: bold;'>
            Объект: {row['Объект']}<br>
            Новость: {row['Заголовок']}<br>
            Тональность: {sentiment}<br>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.write(f"Объект: {row['Объект']}")
        st.write(f"Новость: {row['Заголовок']}")
        st.write(f"Тональность: {sentiment}")
    
    st.write("---")




    
# Initialize sentiment analyzers
finbert = pipeline("sentiment-analysis", model="ProsusAI/finbert")
roberta = pipeline("sentiment-analysis", model="cardiffnlp/twitter-roberta-base-sentiment")
finbert_tone = pipeline("sentiment-analysis", model="yiyanghkust/finbert-tone")


def get_mapped_sentiment(result):
    label = result['label'].lower()
    if label in ["positive", "label_2", "pos", "pos_label"]:
        return "Positive"
    elif label in ["negative", "label_0", "neg", "neg_label"]:
        return "Negative"
    return "Neutral"



def analyze_sentiment(text):
    try:
        finbert_result = get_mapped_sentiment(
            finbert(text, truncation=True, max_length=512)[0]
        )
        roberta_result = get_mapped_sentiment(
            roberta(text, truncation=True, max_length=512)[0]
        )
        finbert_tone_result = get_mapped_sentiment(
            finbert_tone(text, truncation=True, max_length=512)[0]
        )
        
        # Count occurrences of each sentiment
        sentiments = [finbert_result, roberta_result, finbert_tone_result]
        sentiment_counts = {s: sentiments.count(s) for s in set(sentiments)}
        
        # Return sentiment if at least two models agree
        for sentiment, count in sentiment_counts.items():
            if count >= 2:
                return sentiment
                
        # Default to Neutral if no agreement
        return "Neutral"
        
    except Exception as e:
        st.warning(f"Sentiment analysis error: {str(e)}")
        return "Neutral"


def fuzzy_deduplicate(df, column, threshold=50):
    seen_texts = []
    indices_to_keep = []
    for i, text in enumerate(df[column]):
        if pd.isna(text):
            indices_to_keep.append(i)
            continue
        text = str(text)
        if not seen_texts or all(fuzz.ratio(text, seen) < threshold for seen in seen_texts):
            seen_texts.append(text)
            indices_to_keep.append(i)
    return df.iloc[indices_to_keep]


def init_langchain_llm(model_choice):
    try:
        if model_choice == "Qwen2.5-Coder":
            st.info("Loading Qwen2.5-Coder model. только GPU!")
            return QwenSystem()
            
        elif model_choice == "Groq (llama-3.1-70b)":
            if 'groq_key' not in st.secrets:
                st.error("Groq API key not found in secrets. Please add it with the key 'groq_key'.")
                st.stop()
                
            return ChatOpenAI(
                base_url="https://api.groq.com/openai/v1",
                model="llama-3.1-70b-versatile",
                openai_api_key=st.secrets['groq_key'],
                temperature=0.0
            )
            
        elif model_choice == "ChatGPT-4-mini":
            if 'openai_key' not in st.secrets:
                st.error("OpenAI API key not found in secrets. Please add it with the key 'openai_key'.")
                st.stop()
                
            return ChatOpenAI(
                model="gpt-4",
                openai_api_key=st.secrets['openai_key'],
                temperature=0.0
            )
            
        elif model_choice == "Local-MT5":
            return FallbackLLMSystem()
            
    except Exception as e:
        st.error(f"Error initializing the LLM: {str(e)}")
        st.stop()


def estimate_impact(llm, news_text, entity):
    """
    Estimate impact using Groq LLM with improved error handling and validation.
    """
    try:
        # Input validation
        if not news_text or not entity:
            return "Неопределенный эффект", "Недостаточно данных для анализа"
            
        # Clean up inputs
        news_text = str(news_text).strip()
        entity = str(entity).strip()
        
        # Always try to use Groq first
        working_llm = ensure_groq_llm() if 'groq_key' in st.secrets else llm
        
        template = """
        You are a financial analyst tasked with assessing the impact of news on a company.
        
        Company: {entity}
        News Text: {news}
        
        Based on the news content, strictly classify the potential impact into ONE of these categories:
        1. "Значительный риск убытков" - For severe negative events like bankruptcy, major legal issues, significant market loss
        2. "Умеренный риск убытков" - For moderate negative events like minor legal issues, temporary setbacks
        3. "Незначительный риск убытков" - For minor negative events with limited impact
        4. "Вероятность прибыли" - For positive events that could lead to profit or growth
        5. "Неопределенный эффект" - Only if impact cannot be determined from the information
        
        FORMAT YOUR RESPONSE EXACTLY AS:
        Impact: [category name exactly as shown above]
        Reasoning: [2-3 concise sentences explaining your choice]
        """
        
        prompt = PromptTemplate(template=template, input_variables=["entity", "news"])
        chain = prompt | working_llm
        
        # Make the API call
        response = chain.invoke({
            "entity": entity,
            "news": news_text
        })
        
        # Parse response
        response_text = response.content if hasattr(response, 'content') else str(response)
        
        # Extract impact and reasoning
        impact = "Неопределенный эффект"  # Default
        reasoning = "Не удалось определить влияние"  # Default
        
        if "Impact:" in response_text and "Reasoning:" in response_text:
            parts = response_text.split("Reasoning:")
            impact_part = parts[0].split("Impact:")[1].strip()
            reasoning = parts[1].strip()
            
            # Validate impact category with fuzzy matching
            valid_impacts = [
                "Значительный риск убытков",
                "Умеренный риск убытков",
                "Незначительный риск убытков",
                "Вероятность прибыли",
                "Неопределенный эффект"
            ]
            
            # Use fuzzy matching
            best_match = None
            best_score = 0
            for valid_impact in valid_impacts:
                score = fuzz.ratio(impact_part.lower(), valid_impact.lower())
                if score > best_score and score > 80:  # 80% similarity threshold
                    best_score = score
                    best_match = valid_impact
            
            if best_match:
                impact = best_match
        
        return impact, reasoning
        
    except Exception as e:
        st.warning(f"Impact estimation error: {str(e)}")
        if 'rate limit' in str(e).lower():
            st.warning("Rate limit reached. Using fallback analysis.")
        return "Неопределенный эффект", "Ошибка при анализе влияния"

def format_elapsed_time(seconds):
    hours, remainder = divmod(int(seconds), 3600)
    minutes, seconds = divmod(remainder, 60)
    
    time_parts = []
    if hours > 0:
        time_parts.append(f"{hours} час{'ов' if hours != 1 else ''}")
    if minutes > 0:
        time_parts.append(f"{minutes} минут{'' if minutes == 1 else 'ы' if 2 <= minutes <= 4 else ''}")
    if seconds > 0 or not time_parts:
        time_parts.append(f"{seconds} секунд{'а' if seconds == 1 else 'ы' if 2 <= seconds <= 4 else ''}")
    
    return " ".join(time_parts)

def generate_sentiment_visualization(df):
    negative_df = df[df['Sentiment'] == 'Negative']
    
    if negative_df.empty:
        st.warning("Не обнаружено негативных упоминаний. Отображаем общую статистику по объектам.")
        entity_counts = df['Объект'].value_counts()
    else:
        entity_counts = negative_df['Объект'].value_counts()
    
    if len(entity_counts) == 0:
        st.warning("Нет данных для визуализации.")
        return None
    
    fig, ax = plt.subplots(figsize=(12, max(6, len(entity_counts) * 0.5)))
    entity_counts.plot(kind='barh', ax=ax)
    ax.set_title('Количество негативных упоминаний по объектам')
    ax.set_xlabel('Количество упоминаний')
    plt.tight_layout()
    return fig

def create_analysis_data(df):
    analysis_data = []
    for _, row in df.iterrows():
        if row['Sentiment'] == 'Negative':
            analysis_data.append([
                row['Объект'], 
                row['Заголовок'], 
                'РИСК УБЫТКА', 
                row['Impact'],
                row['Reasoning'],
                row['Выдержки из текста']
            ])
    return pd.DataFrame(analysis_data, columns=[
        'Объект', 
        'Заголовок', 
        'Признак', 
        'Оценка влияния',
        'Обоснование',
        'Текст сообщения'
    ])

def translate_reasoning_to_russian(llm, text):
    """Modified to handle both standard LLMs and FallbackLLMSystem"""
    if isinstance(llm, FallbackLLMSystem):
        # Direct translation using MT5
        response = llm.invoke({
            'template_result': f"Translate to Russian: {text}"
        })
        return response.content.strip()
    else:
        # Original LangChain approach
        template = """
        Translate this English explanation to Russian, maintaining a formal business style:
        "{text}"
        
        Your response should contain only the Russian translation.
        """
        prompt = PromptTemplate(template=template, input_variables=["text"])
        chain = prompt | llm
        response = chain.invoke({"text": text})
        
        # Handle different response types
        if hasattr(response, 'content'):
            return response.content.strip()
        elif isinstance(response, str):
            return response.strip()
        else:
            return str(response).strip()


def create_output_file(df, uploaded_file, llm):
    """Create Excel file with multiple sheets from processed DataFrame"""
    try:
        wb = load_workbook("sample_file.xlsx")
        
        # 1. Update 'Публикации' sheet
        ws = wb['Публикации']
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 2. Update 'Мониторинг' sheet with events
        ws = wb['Мониторинг']
        row_idx = 4
        events_df = df[df['Event_Type'] != 'Нет'].copy()
        for _, row in events_df.iterrows():
            ws.cell(row=row_idx, column=5, value=row['Объект'])
            ws.cell(row=row_idx, column=6, value=row['Заголовок'])
            ws.cell(row=row_idx, column=7, value=row['Event_Type'])
            ws.cell(row=row_idx, column=8, value=row['Event_Summary'])
            ws.cell(row=row_idx, column=9, value=row['Выдержки из текста'])
            row_idx += 1

        # 3. Update 'Сводка' sheet
        ws = wb['Сводка']
        entity_stats = pd.DataFrame({
            'Объект': df['Объект'].unique()
        })
        entity_stats['Всего'] = df.groupby('Объект').size()
        entity_stats['Негативные'] = df[df['Sentiment'] == 'Negative'].groupby('Объект').size().fillna(0).astype(int)
        entity_stats['Позитивные'] = df[df['Sentiment'] == 'Positive'].groupby('Объект').size().fillna(0).astype(int)
        
        for idx, (entity, row) in enumerate(entity_stats.iterrows(), start=4):
            ws.cell(row=idx, column=5, value=entity)
            ws.cell(row=idx, column=6, value=row['Всего'])
            ws.cell(row=idx, column=7, value=row['Негативные'])
            ws.cell(row=idx, column=8, value=row['Позитивные'])
            # Get impact for entity
            entity_df = df[df['Объект'] == entity]
            negative_df = entity_df[entity_df['Sentiment'] == 'Negative']
            impact = negative_df['Impact'].iloc[0] if len(negative_df) > 0 else 'Неопределенный эффект'
            ws.cell(row=idx, column=9, value=impact)

        # 4. Update 'Значимые' sheet
        ws = wb['Значимые']
        row_idx = 3
        sentiment_df = df[df['Sentiment'].isin(['Negative', 'Positive'])].copy()
        for _, row in sentiment_df.iterrows():
            ws.cell(row=row_idx, column=3, value=row['Объект'])
            ws.cell(row=row_idx, column=4, value='релевантно')
            ws.cell(row=row_idx, column=5, value=row['Sentiment'])
            ws.cell(row=row_idx, column=6, value=row.get('Impact', 'Неопределенный эффект'))
            ws.cell(row=row_idx, column=7, value=row['Заголовок'])
            ws.cell(row=row_idx, column=8, value=row['Выдержки из текста'])
            row_idx += 1

        # 5. Update 'Анализ' sheet
        ws = wb['Анализ']
        row_idx = 4
        negative_df = df[df['Sentiment'] == 'Negative'].copy()
        for _, row in negative_df.iterrows():
            ws.cell(row=row_idx, column=5, value=row['Объект'])
            ws.cell(row=row_idx, column=6, value=row['Заголовок'])
            ws.cell(row=row_idx, column=7, value="Риск убытка")
            ws.cell(row=row_idx, column=8, value=row.get('Reasoning', 'Не проанализировано'))
            ws.cell(row=row_idx, column=9, value=row['Выдержки из текста'])
            row_idx += 1

        # 6. Update 'Тех.приложение' sheet
        if 'Тех.приложение' not in wb.sheetnames:
            wb.create_sheet('Тех.приложение')
        ws = wb['Тех.приложение']
        
        tech_cols = ['Объект', 'Заголовок', 'Выдержки из текста', 'Translated', 'Sentiment', 'Impact', 'Reasoning']
        tech_df = df[tech_cols].copy()
        
        for r_idx, row in enumerate(dataframe_to_rows(tech_df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Save workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        st.error(f"Error creating output file: {str(e)}")
        st.error(f"DataFrame shape: {df.shape}")
        st.error(f"Available columns: {df.columns.tolist()}")
        return None

def main():
    st.set_page_config(layout="wide")
    
    with st.sidebar:
        st.title("::: AI-анализ мониторинга новостей (v.4.10):::")
        st.subheader("по материалам СКАН-ИНТЕРФАКС")
        
        model_choice = st.radio(
            "Выберите модель для анализа:",
            ["Local-MT5", "Qwen2.5-Coder", "Groq (llama-3.1-70b)", "ChatGPT-4-mini"],
            key="model_selector",
            help="Выберите модель для анализа новостей"
        )
        
        uploaded_file = st.file_uploader(
            "Выбирайте Excel-файл",
            type="xlsx",
            key="file_uploader"
        )
        
        st.markdown(
            """
            Использованы технологии:  
            - Анализ естественного языка с помощью предтренированных нейросетей **BERT**
            - Дополнительная обработка при помощи больших языковых моделей (**LLM**)
            - Фреймворк **LangChain** для оркестрации
            """,
            unsafe_allow_html=True
        )

        st.markdown(
        """
        <style>
        .signature {
            position: fixed;
            right: 12px;
            down: 12px;
            font-size: 14px;
            color: #FF0000;
            opacity: 0.9;
            z-index: 999;
        }
        </style>
        <div class="signature">denis.pokrovsky.npff</div>
        """,
        unsafe_allow_html=True
        )

    # Main content area
    st.title("Анализ мониторинга новостей")
    
    # Initialize session state
    if 'processed_df' not in st.session_state:
        st.session_state.processed_df = None
        
    # Create display areas
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # Area for real-time updates
        st.subheader("Что найдено, показываю:")
        st.markdown("""
            <style>
            .stProgress .st-bo {
                background-color: #f0f2f6;
            }
            .negative-alert {
                background-color: #ffebee;
                border-left: 5px solid #f44336;
                padding: 10px;
                margin: 5px 0;
            }
            .event-alert {
                background-color: #e3f2fd;
                border-left: 5px solid #2196f3;
                padding: 10px;
                margin: 5px 0;
            }
            </style>
        """, unsafe_allow_html=True)
        
    with col2:
        # Area for statistics
        st.subheader("Статистика")
        if st.session_state.processed_df is not None:
            st.metric("Всего статей", len(st.session_state.processed_df))
            st.metric("Из них негативных", 
                len(st.session_state.processed_df[
                    st.session_state.processed_df['Sentiment'] == 'Negative'
                ])
            )
            st.metric("Событий обнаружено", 
                len(st.session_state.processed_df[
                    st.session_state.processed_df['Event_Type'] != 'Нет'
                ])
            )
    
    if uploaded_file is not None and st.session_state.processed_df is None:
        start_time = time.time()
        
        try:
            st.session_state.processed_df = process_file(
                uploaded_file,
                model_choice,
                translation_method='auto'
            )
            
            if st.session_state.processed_df is not None:
                end_time = time.time()
                elapsed_time = format_elapsed_time(end_time - start_time)
                
                # Show results
                st.subheader("Итого по результатам")
                
                # Display statistics
                stats_cols = st.columns(4)
                with stats_cols[0]:
                    st.metric("Всего обработано", len(st.session_state.processed_df))
                with stats_cols[1]:
                    st.metric("Негативных", 
                        len(st.session_state.processed_df[
                            st.session_state.processed_df['Sentiment'] == 'Negative'
                        ])
                    )
                with stats_cols[2]:
                    st.metric("Событий обнаружено", 
                        len(st.session_state.processed_df[
                            st.session_state.processed_df['Event_Type'] != 'Нет'
                        ])
                    )
                with stats_cols[3]:
                    st.metric("Время обработки составило", elapsed_time)
                
                # Show data previews
                with st.expander("📊 Предпросмотр данных", expanded=True):
                    preview_cols = ['Объект', 'Заголовок', 'Sentiment', 'Event_Type']
                    st.dataframe(
                        st.session_state.processed_df[preview_cols],
                        use_container_width=True
                    )
                
                # Create downloadable report
                output = create_output_file(
                    st.session_state.processed_df,
                    uploaded_file,
                    init_langchain_llm(model_choice)
                )
                
                st.download_button(
                    label="📥 Полный отчет - загрузить",
                    data=output,
                    file_name="результаты_анализа.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='download_button'
                )
                
        except Exception as e:
            st.error(f"Ошибочка в обработке файла: {str(e)}")
            st.session_state.processed_df = None


if __name__ == "__main__":
    main()